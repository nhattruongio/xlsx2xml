# -*- encoding: utf-8 -*-
# Copyright (C) 2018

from openpyxl import load_workbook
import random

wb = load_workbook("qp.xlsx")
ws = wb.worksheets[0]

pattern = """
    <record id="{record_id}" model="{model_name}">
        <field name="code">{record_code}</field>
        <field name="name">{record_hospital_name}</field>
        <field name="source">{record_province_id}></field>
    </record>"""
data = random.randint(1, 13000)
model_name = 'hospital'
results = ['''<?xml version='1.0' encoding='utf-8'?>
<odoo>''',
]

# <field name="province_id" model="res.country.state" search="[('code', '=', '{record_province_id}'),('country_id.code', '=', 'VN')]"></field>

for row in ws:
    row = [cell.value for cell in row]

    recode_define = pattern.format(
        record_id='hospital' + row[2] or '',
        model_name=model_name or '',
        record_hospital_name=row[3] or '',
        record_code=row[2] or '',
        record_province_id=row[0] or ''
    )
    results.append(recode_define)
results.append(
'''
</odoo>
'''
)
a = '\n'.join(results)
# print(a)

with open("masterdataqp.xml", "w") as file:
    file.write(str(a))










    # print(doc.getvalue())
# with tag("odoo"):
#     for row in ws.iter_rows(min_row=2, max_row=1300, min_col=1, max_col=10):
#         row = [cell.value for cell in row]
#         with tag("record"):
#             with tag("field"):
#                 text(row[5])
#             with tag("fields"):
#                 text(row[6])
#             # with tag("field"):
#             #     text(row[4])
# result = indent(
#     doc.getvalue(),
#     indentation = '    ',
#     indent_text = True
# )

